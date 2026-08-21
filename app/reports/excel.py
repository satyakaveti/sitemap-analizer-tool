from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import REPORTS_DIR
from app import db

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _score_rating(score: int) -> str:
    if score >= 90:
        return "Excellent"
    elif score >= 75:
        return "Good"
    elif score >= 60:
        return "Needs improvement"
    elif score >= 40:
        return "Poor"
    return "Critical"


def _avg_score(results: list[dict]) -> float:
    scores = [r.get("score", 0) for r in results]
    return round(sum(scores) / len(scores), 1) if scores else 0


def _fmt_issues(issues: list) -> list[str]:
    out = []
    for issue in issues:
        if isinstance(issue, dict):
            code = issue.get("code", "")
            msg = issue.get("message", "")
            out.append(f"[{code}] {msg}" if msg else code)
        else:
            out.append(str(issue))
    return out


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)[:60]))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def _status_fill(code):
    if code is None:
        return None
    if 200 <= code < 300:
        return GREEN_FILL
    if 300 <= code < 400:
        return YELLOW_FILL
    if code >= 400:
        return RED_FILL
    return None


def _write_row(ws, row_num, values, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = fill


def generate_report(scan_id: str) -> str:
    timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H%M%S")
    filename = f"sitemap_scan_{timestamp}_{scan_id}.xlsx"
    filepath = REPORTS_DIR / filename

    status = db.get_status(scan_id)
    results = db.get_results(scan_id)
    error_summary = db.get_error_summary(scan_id)
    seo_summary = db.get_seo_summary(scan_id)
    content_summary = db.get_content_summary(scan_id)

    wb = Workbook()

    _write_summary(wb, status, results, error_summary, seo_summary, content_summary)
    _write_all_urls(wb, results)
    _write_errors(wb, results)
    _write_error_summary(wb, error_summary)
    _write_seo_issues(wb, results, seo_summary)
    _write_content_issues(wb, results, content_summary)
    _write_redirects(wb, results)
    _write_links_images(wb, results)
    _write_duplicates(wb, results)

    wb.save(str(filepath))
    return str(filepath)


def _write_summary(wb, status, results, error_summary, seo_summary, content_summary):
    ws = wb.active
    ws.title = "Summary"
    headers = ["Metric", "Value"]
    ws.append(headers)
    _style_header(ws, len(headers))

    sitemaps = status.get("sitemaps", []) if status else []
    rows = [
        ("Sitemaps", ", ".join(sitemaps) if isinstance(sitemaps, list) else str(sitemaps)),
        ("Date", status.get("started_at", "") if status else ""),
        ("Duration (seconds)", status.get("elapsed", 0) if status else 0),
        ("Total URLs", status.get("total_urls", 0) if status else 0),
        ("", ""),
        ("200 OK", status.get("success", 0) if status else 0),
        ("Redirects", status.get("redirects", 0) if status else 0),
        ("4xx Client Errors", status.get("client_errors", 0) if status else 0),
        ("5xx Server Errors", status.get("server_errors", 0) if status else 0),
        ("Timeouts", status.get("timeouts", 0) if status else 0),
        ("DNS Errors", status.get("dns_errors", 0) if status else 0),
        ("SSL Errors", status.get("ssl_errors", 0) if status else 0),
        ("Other Errors", status.get("other_errors", 0) if status else 0),
        ("", ""),
        ("SEO Issues", status.get("seo_issues", 0) if status else 0),
        ("Content Issues", status.get("content_issues", 0) if status else 0),
        ("", ""),
        ("Average Score", _avg_score(results)),
        ("Excellent (90–100)", sum(1 for r in results if r.get("score", 0) >= 90)),
        ("Good (75–89)", sum(1 for r in results if 75 <= r.get("score", 0) < 90)),
        ("Needs improvement (60–74)", sum(1 for r in results if 60 <= r.get("score", 0) < 75)),
        ("Poor (40–59)", sum(1 for r in results if 40 <= r.get("score", 0) < 60)),
        ("Critical (0–39)", sum(1 for r in results if r.get("score", 0) < 40)),
    ]
    for row in rows:
        ws.append(row)

    ws.append(("", ""))
    ws.append(("Top Error Groups", "Count"))
    _style_header(ws, 2)
    for es in error_summary[:10]:
        ws.append((es["error_type"], es["count"]))

    _auto_width(ws)


def _write_all_urls(wb, results):
    ws = wb.create_sheet("All URLs")
    headers = [
        "URL", "Status", "Score", "Final URL", "Redirect Count", "Response Time",
        "Content Type", "Content Size", "Title", "Title Length",
        "Meta Description", "Meta Desc Length", "H1", "H1 Count",
        "Word Count", "Internal Links", "External Links", "Images",
        "Canonical", "Robots", "Indexable", "Issues",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in results:
        issues = r.get("issues", [])
        if isinstance(issues, str):
            import json
            issues = json.loads(issues) if issues else []
        row = [
            r.get("url"), r.get("status_code"),
            f"{r.get('score', 0)} ({_score_rating(r.get('score', 0))})",
            r.get("final_url"),
            r.get("redirect_count", 0), r.get("response_time", 0),
            r.get("content_type", ""), r.get("content_length", 0),
            r.get("title", ""), r.get("title_length", 0),
            r.get("meta_description", ""), r.get("meta_description_length", 0),
            r.get("h1", ""), r.get("h1_count", 0),
            r.get("word_count", 0),
            r.get("internal_link_count", 0),
            r.get("external_link_count", 0),
            r.get("image_count", 0),
            r.get("canonical", ""),
            r.get("robots", ""), "Yes" if r.get("indexable", True) else "No",
            "; ".join(_fmt_issues(issues)) if issues else "",
        ]
        fill = _status_fill(r.get("status_code"))
        _write_row(ws, ws.max_row + 1, row, fill)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_errors(wb, results):
    ws = wb.create_sheet("Errors")
    headers = ["URL", "Status", "Error", "Response Time"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in results:
        sc = r.get("status_code")
        err = r.get("error", "")
        if (sc and sc >= 400) or err:
            _write_row(ws, ws.max_row + 1, [
                r.get("url"), sc or "N/A", err or "", r.get("response_time", 0)
            ], RED_FILL)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_error_summary(wb, error_summary):
    ws = wb.create_sheet("Error Summary")
    headers = ["Error Type", "Count", "Sample URLs"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for es in error_summary:
        sample = "\n".join(es.get("sample_urls", []))
        _write_row(ws, ws.max_row + 1, [
            es["error_type"], es["count"], sample
        ], RED_FILL if "5" in es.get("error_type", "")[:3] else ORANGE_FILL)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_seo_issues(wb, results, seo_summary):
    ws = wb.create_sheet("SEO Issues")
    headers = ["URL", "Status", "Priority", "Issue Code", "Message"]
    ws.append(headers)
    _style_header(ws, len(headers))

    seo_codes = {
        "TITLE_MISSING", "TITLE_SHORT", "TITLE_LONG", "TITLE_VERY_LONG",
        "META_DESC_MISSING", "META_DESC_SHORT", "META_DESC_LONG", "META_DESC_VERY_LONG",
        "H1_MISSING", "H1_EMPTY", "H1_MULTIPLE",
        "CANONICAL_MISSING", "CANONICAL_CROSS_DOMAIN", "CANONICAL_INVALID", "CANONICAL_MULTIPLE",
        "ROBOTS_NOINDEX", "ROBOTS_NOFOLLOW", "ROBOTS_TXT_BLOCKED",
        "VIEWPORT_MISSING", "OG_MISSING", "OG_PARTIAL",
        "SD_INVALID_JSONLD", "SD_MISSING_PROPERTY", "SD_INVALID_VALUE",
        "URL_HTTP_NOT_HTTPS",
    }
    for r in results:
        issues = r.get("issues", [])
        if isinstance(issues, str):
            import json
            issues = json.loads(issues) if issues else []
        for issue in issues:
            if isinstance(issue, dict):
                code = issue.get("code", "")
                if code in seo_codes:
                    _write_row(ws, ws.max_row + 1, [
                        r.get("url"), r.get("status_code") or "N/A",
                        issue.get("priority", "info"), code, issue.get("message", "")
                    ], ORANGE_FILL)
            else:
                issue_str = str(issue).lower()
                if any(k in issue_str for k in ["title", "meta", "h1", "canonical", "noindex"]):
                    _write_row(ws, ws.max_row + 1, [
                        r.get("url"), r.get("status_code") or "N/A",
                        "medium", "LEGACY", str(issue)
                    ], ORANGE_FILL)

    ws.append(("", ""))
    ws.append(("SEO Issue Summary", "Count"))
    _style_header(ws, 2)
    for s in seo_summary:
        ws.append((s["issue"], s["count"]))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_content_issues(wb, results, content_summary):
    ws = wb.create_sheet("Content Issues")
    headers = ["URL", "Status", "Priority", "Issue Code", "Message"]
    ws.append(headers)
    _style_header(ws, len(headers))

    content_codes = {
        "CONTENT_THIN", "CONTENT_VERY_THIN", "SOFT_404_POSSIBLE", "SOFT_404_LIKELY",
        "SOFT_404_STRONG", "APP_ERROR_PAGE", "CONTENT_WRONG_TYPE",
    }
    for r in results:
        issues = r.get("issues", [])
        if isinstance(issues, str):
            import json
            issues = json.loads(issues) if issues else []
        for issue in issues:
            if isinstance(issue, dict):
                code = issue.get("code", "")
                if code in content_codes:
                    _write_row(ws, ws.max_row + 1, [
                        r.get("url"), r.get("status_code") or "N/A",
                        issue.get("priority", "info"), code, issue.get("message", "")
                    ], ORANGE_FILL)
            else:
                issue_str = str(issue).lower()
                if any(k in issue_str for k in ["thin", "soft", "error", "word"]):
                    _write_row(ws, ws.max_row + 1, [
                        r.get("url"), r.get("status_code") or "N/A",
                        "medium", "LEGACY", str(issue)
                    ], ORANGE_FILL)

    ws.append(("", ""))
    ws.append(("Content Issue Summary", "Count"))
    _style_header(ws, 2)
    for s in content_summary:
        ws.append((s["issue"], s["count"]))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_redirects(wb, results):
    ws = wb.create_sheet("Redirects")
    headers = ["Original URL", "Final URL", "Status", "Redirect Count", "Chain"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in results:
        if r.get("redirect_count", 0) > 0:
            chain = r.get("redirect_chain", [])
            if isinstance(chain, str):
                import json
                chain = json.loads(chain) if chain else []
            final = r.get("final_url", r.get("url", ""))
            chain_str = " -> ".join(chain + [final])
            _write_row(ws, ws.max_row + 1, [
                r.get("url"), final, r.get("status_code"),
                r.get("redirect_count", 0), chain_str
            ], YELLOW_FILL)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_duplicates(wb, results):
    ws = wb.create_sheet("Duplicates")
    headers = ["Issue Type", "Value", "URL"]
    ws.append(headers)
    _style_header(ws, len(headers))

    title_map: dict[str, list[str]] = {}
    desc_map: dict[str, list[str]] = {}
    canonical_map: dict[str, list[str]] = {}
    h1_map: dict[str, list[str]] = {}

    for r in results:
        if r.get("title"):
            title_map.setdefault(r["title"], []).append(r["url"])
        if r.get("meta_description"):
            desc_map.setdefault(r["meta_description"], []).append(r["url"])
        if r.get("canonical"):
            canonical_map.setdefault(r["canonical"], []).append(r["url"])
        if r.get("h1"):
            h1_map.setdefault(r["h1"], []).append(r["url"])

    for label, mapping in [("Duplicate Title", title_map), ("Duplicate Description", desc_map),
                           ("Duplicate Canonical", canonical_map), ("Duplicate H1", h1_map)]:
        for value, urls in mapping.items():
            if len(urls) > 1:
                for url in urls:
                    _write_row(ws, ws.max_row + 1, [label, value[:100], url])

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_links_images(wb, results):
    ws = wb.create_sheet("Links & Images")
    headers = ["URL", "Internal Links", "External Links", "Images", "Images Missing Alt", "Issues"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in results:
        il = r.get("internal_link_count", 0)
        el = r.get("external_link_count", 0)
        ic = r.get("image_count", 0)
        ia = r.get("image_no_alt_count", 0)
        issues = r.get("issues", [])
        if isinstance(issues, str):
            import json
            issues = json.loads(issues) if issues else []
        link_issues = [
            _fmt_issues([i])[0] for i in issues
            if isinstance(i, dict) and i.get("code", "").startswith("LINK_")
        ]
        if il or el or ic:
            _write_row(ws, ws.max_row + 1, [
                r.get("url"), il, el, ic, ia,
                "; ".join(link_issues) if link_issues else "",
            ])

    total_il = sum(r.get("internal_link_count", 0) for r in results)
    total_el = sum(r.get("external_link_count", 0) for r in results)
    total_ic = sum(r.get("image_count", 0) for r in results)
    total_ia = sum(r.get("image_no_alt_count", 0) for r in results)

    ws.append(("", ""))
    ws.append(("Total Internal Links", total_il))
    ws.append(("Total External Links", total_el))
    ws.append(("Total Images", total_ic))
    ws.append(("Total Images Missing Alt", total_ia))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _auto_width(ws)
