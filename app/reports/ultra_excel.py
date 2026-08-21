from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import REPORTS_DIR

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _score_rating(score: int) -> str:
    if score >= 90: return "Excellent"
    if score >= 75: return "Good"
    if score >= 60: return "Needs improvement"
    if score >= 40: return "Poor"
    return "Critical"


def _avg_score(results: list[dict]) -> float:
    scores = [r.get("score", 0) for r in results]
    return round(sum(scores) / len(scores), 1) if scores else 0


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)[:60]))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def _status_fill(code):
    if code is None:
        return None
    if 200 <= code < 300:
        return GREEN_FILL
    if 300 <= code < 400:
        return YELLOW_FILL
    if code >= 400:
        return RED_FILL
    return None


def _write_row(ws, row_num, values, fill=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill:
            cell.fill = fill


def generate_ultra_report(scan_id: str, results: list[dict], elapsed: float, sitemaps: list[str]) -> str:
    timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H%M%S")
    filename = f"sitemap_scan_ultra_{timestamp}_{scan_id}.xlsx"
    filepath = REPORTS_DIR / filename

    wb = Workbook()

    # 1. Summary Sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Metric", "Value"])
    _style_header(ws_sum, 2)

    success = sum(1 for r in results if r.get("status_code") == 200)
    redirects = sum(1 for r in results if r.get("redirect_count", 0) > 0)
    client_err = sum(1 for r in results if r.get("status_code") and 400 <= r.get("status_code") < 500)
    server_err = sum(1 for r in results if r.get("status_code") and r.get("status_code") >= 500)
    timeouts = sum(1 for r in results if r.get("error") == "timeout")
    other_err = len(results) - (success + redirects + client_err + server_err + timeouts)

    rows = [
        ("Scan Type", "Ultra Short Scan (In-Memory)"),
        ("Sitemaps", ", ".join(sitemaps)),
        ("Duration (seconds)", f"{elapsed:.2f}s"),
        ("Total URLs Checked", len(results)),
        ("", ""),
        ("200 OK Success", success),
        ("Redirects", redirects),
        ("4xx Client Errors", client_err),
        ("5xx Server Errors", server_err),
        ("Timeouts", timeouts),
        ("Other Errors", other_err),
        ("", ""),
        ("Average Score", _avg_score(results)),
    ]
    for row in rows:
        ws_sum.append(row)
    _auto_width(ws_sum)

    # 2. All URLs Sheet
    ws_urls = wb.create_sheet("All URLs")
    headers = [
        "URL", "Status Code", "Score", "Final URL", "Response Time (s)",
        "Content Size (Bytes)", "Title", "Word Count", "Error Message", "Issues"
    ]
    ws_urls.append(headers)
    _style_header(ws_urls, len(headers))

    for r in results:
        issues_list = r.get("issues", [])
        issues_str = "; ".join([i.get("message", i.get("code", "")) if isinstance(i, dict) else str(i) for i in issues_list])
        row_val = [
            r.get("url"),
            r.get("status_code") or "N/A",
            f"{r.get('score', 0)} ({_score_rating(r.get('score', 0))})",
            r.get("final_url"),
            r.get("response_time", 0),
            r.get("content_length", 0),
            r.get("title", ""),
            r.get("word_count", 0),
            r.get("error") or "",
            issues_str
        ]
        fill = _status_fill(r.get("status_code"))
        _write_row(ws_urls, ws_urls.max_row + 1, row_val, fill)
    
    ws_urls.auto_filter.ref = ws_urls.dimensions
    ws_urls.freeze_panes = "A2"
    _auto_width(ws_urls)

    wb.save(str(filepath))
    return str(filepath)
